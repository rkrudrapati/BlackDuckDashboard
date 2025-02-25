"""
This script will take input as source directory and destination directory. It is given in the sript itself (source variable: pat, destination: dest)
Inside this path, it will walk through subfolders and try to find the security_*.csv file and format the report.
The reports will be saved in the destination folder
"""
import csv
from openpyxl import Workbook
from openpyxl import styles
try:
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter


def as_text(value):
    if value is None:
        return ""
    return str(value)


def cell_format_style(_cell, header=False):
    side = styles.Side(style='thin')
    temp = _cell.value
    if str(temp).__len__() > 100:
        _cell.alignment = styles.Alignment(horizontal='left', vertical="top", wrap_text=True)
    else:
        _cell.alignment = styles.Alignment(horizontal='center', vertical="top")
    fill_color = styles.Color(rgb='006666FF', tint=0.5)
    _cell.border = styles.Border(top=side, right=side, bottom=side, left=side)
    if header == True:
        _cell.fill = styles.PatternFill(patternType='solid', fgColor=fill_color)


def main_program(source, destination):
    wb = Workbook()
    ws = wb.active
    ws.title = 'sheet 1'

    with open(source, encoding="utf8") as csvfile:
        readCSV = csv.reader(csvfile, delimiter=',')
        _row = 1
        header = True
        for _readline in readCSV:
            if 'components' in source:
                rearranged_data = [_row - 1, _readline[3], _readline[4], _readline[5], _readline[7], _readline[26],
                                   _readline[13], _readline[8], _readline[9], _readline[12], _readline[14],
                                   _readline[15], _readline[16], _readline[17], _readline[18], _readline[19],
                                   _readline[20], _readline[21], _readline[22], _readline[23], _readline[24],
                                   _readline[25], _readline[27], _readline[28], _readline[29], _readline[30],
                                   _readline[31], _readline[32], _readline[33], _readline[34], _readline[35],
                                   _readline[36], _readline[37], _readline[38], _readline[39], _readline[40],
                                   _readline[41], _readline[42], _readline[43]
                                   ]
            for _coln in range(1, rearranged_data.__len__() + 1):
                if rearranged_data[0] == 0:
                    rearranged_data[0] = 'S No'
                ws.cell(row=_row, column=_coln).value = rearranged_data[_coln - 1]
                cell_format_style(ws.cell(row=_row, column=_coln), header)
            header = False
            _row += 1
    wb.save(destination)
    csvfile.close()

    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells) + 2
        if length > 150:
            length = 150
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length
    wb.save(destination)
    wb.close()


pat = r"C:\Users\code1\Desktop\_temp\Maheshan\Docker_Base_Image_Scan-alpine-3.13_2021-05-05_093816"
dest = r"C:\Users\code1\Desktop\_temp\Maheshan\Docker_Base_Image_Scan-alpine-3.13_2021-05-05_093816\License"
from os import walk, path
for root, dirs, files in walk(pat):
    # print(path.join(root, files))
    for file in files:
        if "components_" in file:
            source = path.join(root, file)
            directory = source.split("\\")
            directory = directory[7].strip("Docker_Base_Image_Scan-")
            destination = dest + "\\" + directory + "_license.xlsx"
            print(source)
            main_program(source, destination)
            print(destination)

print('Updated file %s' %dest)