"""
vuln_kpi_report_check_count_matching_or_not is the base.
This script will identify where a particular license is used and back tracks the component/version & project/version details.
In some cases, the component version is not identified by the BD tool. So back tracking is not possible. This means that there is error margin in this script.
The unknown versions are marked NA in the report(excel)
"""

import json
from sys import exit
from re import match
from requests import request
from icecream import ic
from openpyxl import Workbook
from openpyxl import styles

try:
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter

ic.disable()
# server_url = 'https://ENTER_SERVER_URL/'
# api_token = "YzljZjFhYjktMDJlhYmVmLTk3MTEyZjg5ODQ4Ng=="  # Enter api token here
server_url = 'https://blackduckweb.philips.com/'
api_token = "YzljZjFhYjktMDJlhYmVmLTk3MTEyZjg5ODQ4Ng=="  # Enter api token here
if server_url[-1] == "/":
    server_url = server_url[:-1]
    # ic(server_url)
destination = r"C:\Temp\auto_vuln_data.xlsx"


def as_text(value):
    if value is None:
        return ""
    return str(value)


def cell_format_style(_cell, header_frame=False):
    side = styles.Side(style='thin')
    temp = _cell.value
    if str(temp).__len__() > 100:
        _cell.alignment = styles.Alignment(horizontal='left', vertical="top", wrap_text=True)
    else:
        _cell.alignment = styles.Alignment(horizontal='center', vertical="top")
    fill_color = styles.Color(rgb='006666FF', tint=0.5)
    _cell.border = styles.Border(top=side, right=side, bottom=side, left=side)
    if header_frame:  # true
        _cell.fill = styles.PatternFill(patternType='solid', fgColor=fill_color)


def post_request(url, payload, headers):
    try:
        post_response = request("POST", url, data=payload, headers=headers, verify=False)
        if match('20[0-4]', str(post_response.status_code)):
            return post_response
        else:
            print('request failed')
            print('Response status: %d' % post_response.status_code)
            print('Method: post_request')
            print('request response: \n %s' % post_response.text)
            exit()
    except Exception as err:
        print(err)
        exit()


def get_request(url, headers):
    try:
        get_response = request("GET", url, headers=headers, verify=False)
        if match('20[0-4]', str(get_response.status_code)):
            return get_response
        else:
            print('request failed')
            print('Response status: %d' % get_response.status_code)
            print('Method: get_request')
            print('request response: \n %s' % get_response.text)
            exit()
    except Exception as err:
        print(err)
        exit()


###########################
with open('input', 'r', encoding='UTF-8') as file_input:
    # print(file_input)
    json_input = json.load(file_input)
    list_input = []
    for each_item in json_input["items"]:
        # print(f"License Name: {each_item['name']}")
        # print(f"Link: {each_item['_meta']['href']}")
        list_input.append([each_item['name'], each_item['_meta']['href']])
    # for items in list_input:
    #     print(items)
print(list_input)
###########################


wb = Workbook()
ws = wb.active
ws.title = 'vuln_data'
_row = 1
destination = r"C:\Temp\lgpl_license.xlsx"

csrf_token = ''
auth_token = ''
login_headers = {
    'Content-Type': "application/x-www-form-urlencoded; charset=UTF-8",
    'Authorization': 'token {}'.format(api_token),
}

login_url = server_url + '/api/tokens/authenticate'
response = post_request(url=login_url, payload={}, headers=login_headers)
auth_token = response.json()['bearerToken']
csrf_token = response.headers['X-CSRF-TOKEN']
get_headers = {
    'Content-Type': "application/json",
    'cache-control': "no-cache",
    'X-Requested-With': "XMLHttpRequest",
    'Accept': "application/json",
    'X-CSRF-TOKEN': csrf_token,
    'authorization': 'bearer {}'.format(auth_token)
}

"""
https://blackduckweb.philips.com/api/internal/composite/licenses?offset=0&limit=100&sort=&filter=inUse%3Atrue&q=name%3Alesser
"""
for each_item_list_input in list_input:
    license_name = each_item_list_input[0]
    license_href_url = each_item_list_input[1] + "/usages?limit=9999"
    license_href_response = get_request(url=license_href_url, headers=get_headers)
    # print(response.json())
    try:
        for each_items_license_response in license_href_response.json()["items"]:
            login_url = server_url + '/api/tokens/authenticate'
            response = post_request(url=login_url, payload={}, headers=login_headers)
            auth_token = response.json()['bearerToken']
            csrf_token = response.headers['X-CSRF-TOKEN']
            get_headers = {
                'Content-Type': "application/json",
                'cache-control': "no-cache",
                'X-Requested-With': "XMLHttpRequest",
                'Accept': "application/json",
                'X-CSRF-TOKEN': csrf_token,
                'authorization': 'bearer {}'.format(auth_token)
            }
            componentName = each_items_license_response["componentName"]
            try:
                componentVersionName = str(each_items_license_response["componentVersionName"])
            except Exception as err3:  # handling the unknown component version here
                componentVersionName = "Unknown"
                print(f"NA|NA|{componentName}|{componentVersionName}|{license_name}")
                output.append("NA")
                output.append("NA")
                output.append(componentName)
                output.append(componentVersionName)
                output.append(license_name)
                continue
            componentVersion_href = each_items_license_response["_meta"]['links'][1]["href"] + "/references?limit=9999"
            component_version_response = get_request(url=componentVersion_href, headers=get_headers)
            try:
                for each_item_component_version_response in component_version_response.json()["items"]:
                    output = []
                    project_name = each_item_component_version_response["projectName"]
                    if project_name in ["BDvsWS", "SCOE_Testing", "test_project", "test_project_3"]:
                        continue
                    versionName = str(each_item_component_version_response["versionName"])
                    print(f"{project_name}|{versionName}|{componentName}|{componentVersionName}|{license_name}")
                    output.append(project_name)
                    output.append(versionName)
                    output.append(componentName)
                    output.append(componentVersionName)
                    output.append(license_name)
                    # print(output)
                    for _coln in range(1, output.__len__() + 1):
                        ws.cell(row=_row, column=_coln).value = output[_coln - 1]
                        header = False
                        cell_format_style(ws.cell(row=_row, column=_coln), header)
                    _row += 1
                    wb.save(destination)
            except Exception as err2:
                print(err2)
                print("component_version_response level work")
                print(componentName)
                print(componentVersionName)
                print(component_version_response.json())
                print(project_name)
                print(versionName)
                continue
    except Exception as err4:
        print(err4)
        print(license_name)
        print(license_href_response.json())
        print(componentName)
        print(componentVersionName)
        continue
for column_cells in ws.columns:
    length = max(len(str(cell.value)) for cell in column_cells) + 2
    if length > 150:
        length = 150
    ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length
wb.save(destination)
wb.close()
