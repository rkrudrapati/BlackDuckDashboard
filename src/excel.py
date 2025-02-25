from openpyxl import load_workbook
from shutil import copy


def write_to_data_table(lv_BG_name, lv_BIU_list, lv_row_count):
    severity = ["High", "Medium", "Low"]
    for i in range(1,4):
        table_data_sheet.cell(row=lv_row_count+i+1,column=1).value = lv_BG_name
        table_data_sheet.cell(row=lv_row_count+i+1,column=2).value = severity[i-1]

    coln = 3
    for each_BIU_Name in lv_BIU_list:
        formula_high = "=SUMIF('Black Duck Latest Version'!B:B,\"%s\",'Black Duck Latest Version'!G:G)" % each_BIU_Name
        formula_medium = "=SUMIF('Black Duck Latest Version'!B:B,\"%s\",'Black Duck Latest Version'!H:H)" % each_BIU_Name
        formula_low = "=SUMIF('Black Duck Latest Version'!B:B,\"%s\",'Black Duck Latest Version'!I:I)" % each_BIU_Name
        formula = [each_BIU_Name, formula_high, formula_medium, formula_low]
        #table_data_sheet.cell(row=lv_row_count + i, column=coln).value = each_BIU_Name
        for i in range(4):
            table_data_sheet.cell(row=lv_row_count + i+1, column=coln).value = formula[i]
        coln += 1

copy('Template_Design_Updated.xlsx', 'temp.xlsx')
wb = load_workbook(filename="result.xlsx", data_only=True)
#print(wb.sheetnames)
table_data_sheet = wb["Vulnerability_Table"]
latest_version_security_vuln_info_sheet = wb["Black Duck Latest Version"]

i = 0
BG_list = []
for rows in latest_version_security_vuln_info_sheet.rows:
    if rows[0].value != "BG":
        BG_list.append(rows[0].value)
BG_list = sorted(set(BG_list))
#print(BG_list)
row_count = 0
for each_BG in BG_list:
    BIU_list = []
    #print(each_BG)
    for rows in latest_version_security_vuln_info_sheet.rows:
        if rows[0].value == each_BG:
            BIU_list.append(rows[1].value)
        BIU_list = sorted(set(BIU_list))
    # print(BIU_list, end="\t")
    # print("")
    write_to_data_table(each_BG, BIU_list, row_count)
    row_count += 5

wb.save("VulnerabilityTable.xlsx")
wb.close()
