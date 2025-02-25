"""
P1_dormant_projects is the base.
This script will identify the latest or last scanned date of a project.
The script will verify the code locations and their respective scan time
Based on this, it will give the last scanned date of a project
"""

from sys import exit
from re import match
from typing import List, Any

from requests import request
from src.P1_my_requests import get_request
from icecream import ic
from openpyxl import Workbook
from openpyxl import styles
from os import walk
from datetime import date

try:
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter

ic.disable()
# server_url = 'https://ENTER_SERVER_URL/'
# api_token = "YzljZjFhYjktMDJlhYmVmLTk3MTEyZjg5ODQ4Ng=="  # Enter api token here
server_url = 'https://blackduckweb.philips.com/'
api_token = "YzljZjFhYjktMDJlhYmVmLTk3MTEyZjg5ODQ4Ng=="  # Enter api token here
if server_url[-1] == "/":
    server_url = server_url[:-1]
    # ic(server_url)
destination = r"C:\Temp\auto_vuln_data.xlsx"


def as_text(value):
    if value is None:
        return ""
    return str(value)


def cell_format_style(_cell, header=False):
    side = styles.Side(style='thin')
    temp = _cell.value
    if str(temp).__len__() > 100:
        _cell.alignment = styles.Alignment(horizontal='left', vertical="top", wrap_text=True)
    else:
        _cell.alignment = styles.Alignment(horizontal='center', vertical="top")
    fill_color = styles.Color(rgb='006666FF', tint=0.5)
    _cell.border = styles.Border(top=side, right=side, bottom=side, left=side)
    if header:  # true
        _cell.fill = styles.PatternFill(patternType='solid', fgColor=fill_color)


def post_request(url, payload, headers):
    try:
        post_response = request("POST", url, data=payload, headers=headers, verify=False)
        if match('20[0-4]', str(post_response.status_code)):
            return post_response
        else:
            print('request failed')
            print('Response status: %d' % post_response.status_code)
            print('Method: post_request')
            print('request response: \n %s' % post_response.text)
            exit()
    except Exception as err:
        print(err)
        exit()


def get_request(url, headers):
    try:
        get_response = request("GET", url, headers=headers, verify=False)
        if match('20[0-4]', str(get_response.status_code)):
            return get_response
        else:
            print('request failed')
            print('Response status: %d' % get_response.status_code)
            print('Method: get_request')
            print('request response: \n %s' % get_response.text)
            exit()
    except Exception as err:
        print(err)
        exit()


def find_latest_version(lv_project_link):
    lv_versions_link = lv_project_link + "/versions?limit=999"
    lv_versions_request_response = get_request(url=lv_versions_link, headers=get_headers)
    ref_date = "2017-01-01"
    lv_latest_version_name = ""
    lv_latest_version_url = ""
    for v_items in lv_versions_request_response.json()["items"]:
        lv_version_name = v_items['versionName']
        # createdAt = v_items['createdAt'].split('T')[0]  # date
        each_version_link = v_items["_meta"]["href"]
        codelocation_link = each_version_link + "/codelocations"
        codelocation_response = get_request(url=codelocation_link, headers=get_headers)
        count = codelocation_response.json()["totalCount"]
        if count == 0:
            lv_latest_version_name = lv_version_name
            continue
        else:
            for each_code_location_details in codelocation_response.json()["items"]:
                updatedAt = each_code_location_details["updatedAt"].split('T')[0]
                if ref_date < updatedAt:
                    ref_date = updatedAt
                    lv_latest_version_name = lv_version_name
                    lv_latest_version_url = each_version_link
    return [lv_latest_version_name, lv_latest_version_url, ref_date]


csrf_token = ''
auth_token = ''
login_headers = {
    'Content-Type': "application/x-www-form-urlencoded; charset=UTF-8",
    'Authorization': 'token {}'.format(api_token),
}

login_url = server_url + '/api/tokens/authenticate'
response = post_request(url=login_url, payload={}, headers=login_headers)
auth_token = response.json()['bearerToken']
csrf_token = response.headers['X-CSRF-TOKEN']
get_headers = {
    'Content-Type': "application/json",
    'cache-control': "no-cache",
    'X-Requested-With': "XMLHttpRequest",
    'Accept': "application/json",
    'X-CSRF-TOKEN': csrf_token,
    'authorization': 'bearer {}'.format(auth_token)
}

reference_date = "2017-01-01"
today = date.today()
wb = Workbook()
ws = wb.active
ws.title = 'vuln_data'
_row = 1

projects_url = server_url + "/api/projects?limit=999"  # offset=6&limit=1"
project_request_response = get_request(url=projects_url, headers=get_headers)
# print(f'''Project_Name|Version_Name|last scanned date|critical_components_count|high_components_count|medium_components_count|low_components_count|critical_issues_count|high_issues_count|medium_issues_count|low_issues_count|Age|Level''')
my_doc_header = ['Project_Name', 'Version_Name', 'last scanned date', 'critical_components_count',
                 'high_components_count', 'medium_components_count', 'low_components_count', 'critical_issues_count',
                 'high_issues_count', 'medium_issues_count', 'low_issues_count', 'Age', 'Level']
for _coln in range(1, my_doc_header.__len__() + 1):
    ws.cell(row=_row, column=_coln).value = my_doc_header[_coln - 1]
    header = True
    cell_format_style(ws.cell(row=_row, column=_coln), header)
_row += 1
for each_project_request_response_items in project_request_response.json()["items"]:
    project_name = each_project_request_response_items["name"]
    # version_name, latest_version_url, last_updated_date = find_latest_version(each_project_request_response_items["_meta"]["href"])
    version_url = each_project_request_response_items["_meta"]["href"] + "/versions?limit=999"
    version_request_response = get_request(url=version_url, headers=get_headers)
    for each_version_request_response_items in version_request_response.json()["items"]:
        output: list[Any] = [project_name]
        version_name = each_version_request_response_items["versionName"]
        output.append(version_name)
        each_version_url = each_version_request_response_items["_meta"]["href"]
        critical_components_count = 0
        high_components_count = 0
        medium_components_count = 0
        low_components_count = 0
        critical_issues_count = 0
        high_issues_count = 0
        medium_issues_count = 0
        low_issues_count = 0
        Level = ""
        """processing codelocation links to get the scanned date"""
        codelocation_link = each_version_url + "/codelocations"
        codelocation_response = get_request(url=codelocation_link, headers=get_headers)
        count = codelocation_response.json()["totalCount"]
        if count == 0:
            updatedAt = reference_date
            continue
        else:
            for each_code_location_details in codelocation_response.json()["items"]:
                updatedAt = each_code_location_details["updatedAt"].split('T')[0]
        dt = updatedAt.split("-")
        updatedAt = date(int(dt[0]), int(dt[1]), int(dt[2]))
        output.append(updatedAt)
        if today == updatedAt:
            age = 1
        else:
            age = today - updatedAt
            age = age.days
        if age <= 180:
            Level = "Level 3"
        elif 180 < age <= 365:
            Level = "Level 2"
        elif age > 365:
            Level = "Level 1"
        risk_profile_url = each_version_url + "/risk-profile"
        risk_profile_request_response = get_request(url=risk_profile_url, headers=get_headers)
        critical_components_count = risk_profile_request_response.json()["categories"]["VULNERABILITY"]["CRITICAL"]
        high_components_count = risk_profile_request_response.json()["categories"]["VULNERABILITY"]["HIGH"]
        medium_components_count = risk_profile_request_response.json()["categories"]["VULNERABILITY"]["MEDIUM"]
        low_components_count = risk_profile_request_response.json()["categories"]["VULNERABILITY"]["LOW"]
        vuln_bom_url = each_version_url + \
                       "/components?limit=99999&filter=securityRisk%3Acritical&filter=securityRisk%3Ahigh&filter=securityRisk%3Amedium&filter=securityRisk%3Alow"
        vuln_bom_request_response = get_request(url=vuln_bom_url, headers=get_headers)
        for each_vuln_bom_request_response_items in vuln_bom_request_response.json()["items"]:
            for each_item_count in each_vuln_bom_request_response_items["securityRiskProfile"]["counts"]:
                if each_item_count["countType"] == "CRITICAL":
                    critical_issues_count += each_item_count["count"]
                if each_item_count["countType"] == "HIGH":
                    high_issues_count += each_item_count["count"]
                if each_item_count["countType"] == "MEDIUM":
                    medium_issues_count += each_item_count["count"]
                if each_item_count["countType"] == "LOW":
                    low_issues_count += each_item_count["count"]
        temp_lv = [critical_components_count, high_components_count, medium_components_count, low_components_count,
                   critical_issues_count, high_issues_count, medium_issues_count, low_issues_count, age, Level]
        for items in temp_lv:
            output.append(items)
        print(output)
        for _coln in range(1, output.__len__() + 1):
            ws.cell(row=_row, column=_coln).value = output[_coln - 1]
            header = False
            cell_format_style(ws.cell(row=_row, column=_coln), header)
        _row += 1
        wb.save(destination)

for column_cells in ws.columns:
    length = max(len(str(cell.value)) for cell in column_cells) + 2
    if length > 150:
        length = 150
    ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length
wb.save(destination)
wb.close()
# print(f'''{project_name}|{version_name}|{updatedAt}|{critical_components_count}|{high_components_count}|{medium_components_count}|{low_components_count}|{critical_issues_count}|{high_issues_count}|{medium_issues_count}|{low_issues_count}|{age}|{Level}''')
