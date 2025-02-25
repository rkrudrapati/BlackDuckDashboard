"""legacy copy: docker_components_compare.py"""
#  Copyright (c) 2019. Lorem ipsum dolor sit amet, consectetur adipiscing elit.
#  Morbi non lorem porttitor neque feugiat blandit. Ut vitae ipsum eget quam lacinia accumsan.
#  Etiam sed turpis ac ipsum condimentum fringilla. Maecenas magna.
#  Proin dapibus sapien vel ante. Aliquam erat volutpat. Pellentesque sagittis ligula eget metus.
#  Vestibulum commodo. Ut rhoncus gravida arcu.

import csv
from openpyxl import Workbook
from openpyxl import styles
try:
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter

"""
user should define the below variables.
source_csv_path should be a csv file(security.csv). Please give complete path along with the file name.
destination_excel_path should be a excel file(test01.xlsx). Please give complete path along with the desired file name
For successfully running this script, 
1. make sure that source_csv_path is pointing to proper genuine file.
2. make sure the output file is closed(if it already exists)
"""
source_csv_path = r''            # eg: /phoenix_phoenix-2.1.0.0/security.csv
destination_excel_path = r''     # eg: C:/Users/Desktop/Docker_Base_Image_Scan/New_Temp01.xlsx


def as_text(value):
    if value is None:
        return ""
    return str(value)


def cell_format_style(_cell, header=False):
    side = styles.Side(style='thin')
    temp = _cell.value
    if str(temp).__len__() > 100:
        _cell.alignment = styles.Alignment(horizontal='left', vertical="top", wrap_text=True)
    else:
        _cell.alignment = styles.Alignment(horizontal='center', vertical="top")
    fill_color = styles.Color(rgb='006666FF', tint=0.5)
    _cell.border = styles.Border(top=side, right=side, bottom=side, left=side)
    if header == True:
        _cell.fill = styles.PatternFill(patternType='solid', fgColor=fill_color)


wb = Workbook()
ws = wb.active
ws.title = 'security'


with open(source_csv_path) as csvfile:
    readCSV = csv.reader(csvfile, delimiter=',')
    _row = 1
    header = True
    for _readline in readCSV:
        rearranged_data = [_row-1, _readline[4], _readline[5], _readline[6], _readline[10], _readline[22],
                           _readline[17], _readline[16], _readline[9], _readline[21], _readline[23], _readline[1],
                           _readline[2], _readline[3], _readline[7], _readline[8], _readline[13], _readline[14],
                           _readline[15], _readline[19], _readline[20]]
        for _coln in range(1, rearranged_data.__len__()+1):
            if rearranged_data[0] == 0:
                rearranged_data[0] = 'S No'
            ws.cell(row=_row, column=_coln).value = rearranged_data[_coln-1]
            cell_format_style(ws.cell(row=_row, column=_coln), header)
        header = False
        _row += 1
wb.save(destination_excel_path)
csvfile.close()


for column_cells in ws.columns:
    # cell_length = []
    # for cell in column_cells:
    #     cell_length.append(len(str(cell.value)))
    # length = max(cell_length)+2
    length = max(len(str(cell.value)) for cell in column_cells) + 2
    if length > 150:
        length = 150
    ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length


wb.save(destination_excel_path)
wb.close()
